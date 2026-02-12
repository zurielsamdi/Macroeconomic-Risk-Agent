from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from typing import Any, List, Optional

from openpyxl.workbook.workbook import Workbook


@dataclass
class MacroEvent:
    date: str
    region: str
    event: str
    time: str
    risk: str
    importance: int = 0
    expected: Optional[float] = None
    actual: Optional[float] = None
    prior: Optional[float] = None
    unit: Optional[str] = None


def _to_bool(v) -> bool:
    # Already a real boolean
    if isinstance(v, bool):
        return v

    if v is None:
        return False

    # Numbers
    if isinstance(v, (int, float)):
        return v != 0

    # Strings (including Excel formulas like "=TRUE")
    if isinstance(v, str):
        s = v.strip()
        if s.startswith("="):
            s = s[1:].strip()   # "=TRUE" -> "TRUE"
        s = s.upper()

        if s in {"TRUE", "T", "YES", "Y", "1"}:
            return True
        if s in {"FALSE", "F", "NO", "N", "0", ""}:
            return False

    # Fallback
    return False

def _to_int(x):
    try:
        if x is None or str(x).strip() == "":
            return None
        return int(float(str(x).strip()))
    except Exception:
        return None

def _to_float(x):
    try:
        if x is None or str(x).strip() == "":
            return None
        return float(str(x).strip())
    except Exception:
        return None


def load_macro_events_from_wb(wb: Workbook, today: date, sheet_name: str = "Macro_Events") -> List[MacroEvent]:
    """
    Reads events from Macro_Events sheet.
    Filters:
      - Active == TRUE
      - If Date is filled, keep only rows where Date == today's YYYY-MM-DD
      - If Date is blank, treat as 'always show'
    """
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]

    headers = [c.value for c in ws[1]]
    if not headers or all(h is None for h in headers):
        return []

    idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    required = ["Region", "Event", "Time", "Risk", "Active", "Importance"]
    if not all(r in idx for r in required):
        raise ValueError(f"{sheet_name} missing required headers. Need: {required}. Found: {headers}")

    today_str = today.strftime("%Y-%m-%d")
    has_date = "Date" in idx

    out: List[MacroEvent] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        active = _to_bool(row[idx["Active"]])
        if not active:
            continue

        row_date_str = today.strftime("%Y-%m-%d")

        # Date filter (robust)
        if has_date:
            v = row[idx["Date"]]

            if v is None:
                continue

            # Excel often gives datetime.datetime
            if hasattr(v, "date"):  # datetime or date-like
                row_date = v.date() if hasattr(v, "date") else v
            else:
                # fallback: string like "2026-02-07"
                try:
                    row_date = datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
                except Exception:
                    continue
        
            row_date_str = row_date.strftime("%Y-%m-%d")

            if row_date != today:
                continue


        region = str(row[idx["Region"]] or "").strip()
        event = str(row[idx["Event"]] or "").strip()
        time = str(row[idx["Time"]] or "").strip()
        risk = str(row[idx["Risk"]] or "").strip()
        importance = _to_int(row[idx["Importance"]]) if "Importance" in idx else None
        expected = _to_float(row[idx["Expected"]]) if "Expected" in idx else None
        actual = _to_float(row[idx["Actual"]]) if "Actual" in idx else None
        prior = _to_float(row[idx["Prior"]]) if "Prior" in idx else None
        unit = str(row[idx["Unit"]]).strip() if ("Unit" in idx and row[idx["Unit"]] is not None) else None



        if not region or not event:
            continue


        out.append(MacroEvent(row_date_str, region, event, time, risk,
                     importance=importance,
                     expected=expected,
                     actual=actual,
                     prior=prior,
                     unit=unit))

    return out
