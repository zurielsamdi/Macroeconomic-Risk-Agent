from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional, Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# -----------------------------
# Consistent styles (basic)
# -----------------------------
THIN_SIDE = Side(style="thin", color="D9D9D9")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)


@dataclass(frozen=True)
class TableSpec:
    """
    Defines where and how to write a table.
    start_row/start_col are 1-indexed (Excel style).
    """
    start_row: int
    start_col: int
    header: bool = True


def open_template(template_path: Path):
    """
    Load template workbook into memory.
    We never write into the template directly.
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Missing template: {template_path.resolve()}")
    return load_workbook(template_path)


def ensure_sheet(wb, sheet_name: str) -> Worksheet:
    """
    Fail loudly if a required sheet is missing.
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook missing sheet '{sheet_name}'. Found: {wb.sheetnames}")
    return wb[sheet_name]


def write_text(ws: Worksheet, cell: str, value: Any, bold: bool = False):
    """
    Write a single value into a single cell.
    """
    safe_value = sanitize_excel_text(value) if isinstance(value, str) or value is None else value
    ws[cell] = safe_value
    # Guard assertion: auto-escape string formulas if any slipped through.
    if isinstance(ws[cell].value, str) and ws[cell].value[:1] in {"=", "+", "-", "@"}:
        ws[cell] = "'" + ws[cell].value
    ws[cell].alignment = CELL_ALIGN
    if bold:
        ws[cell].font = Font(bold=True)


def write_bullets(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    title: str,
    lines: Sequence[str],
) -> int:
    """
    Write a titled bullet block and return the last row written.
    """
    col = get_column_letter(start_col)
    write_text(ws, f"{col}{start_row}", title, bold=True)
    r = start_row + 1
    for line in lines:
        write_text(ws, f"{col}{r}", line)
        r += 1
    return r - 1


def clear_range(ws: Worksheet, min_row: int, min_col: int, max_row: int, max_col: int):
    """
    Clear values in a rectangular region.
    Prevents leftover data from prior runs.
    """
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).value = None


def write_table(
    ws: Worksheet,
    spec: TableSpec,
    columns: Sequence[str],
    rows: Sequence[Sequence[Any]],
    number_formats: Optional[Sequence[Optional[str]]] = None,
    max_clear_rows: int = 200,
):
    """
    Write a rectangular table with optional header + number formats.
    """
    n_cols = len(columns)
    start_r, start_c = spec.start_row, spec.start_col

    # Clear a safe region first
    clear_range(
        ws,
        min_row=start_r,
        min_col=start_c,
        max_row=start_r + max_clear_rows,
        max_col=start_c + max(0, n_cols - 1),
    )

    write_r = start_r

    # Header row
    if spec.header:
        for j, col_name in enumerate(columns):
            cell = ws.cell(row=write_r, column=start_c + j, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        write_r += 1

    # Body
    for i, row in enumerate(rows):
        if len(row) != n_cols:
            raise ValueError(f"Row {i} length {len(row)} != expected {n_cols}")

        for j, val in enumerate(row):
            if isinstance(val, str) or val is None:
                safe_val = sanitize_excel_text(val)
            else:
                safe_val = val
            cell = ws.cell(row=write_r + i, column=start_c + j, value=safe_val)
            # Guard assertion: auto-escape any leading formula markers in strings.
            if isinstance(cell.value, str) and cell.value[:1] in {"=", "+", "-", "@"}:
                cell.value = "'" + cell.value
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

            if number_formats and number_formats[j]:
                cell.number_format = number_formats[j]


def sanitize_excel_text(value: Any, max_len: int = 30000) -> str:
    """
    Sanitize freeform text before writing to Excel cells.
    - None -> ""
    - Replace CR/LF/TAB with spaces
    - Strip ASCII controls 0x00-0x1F except space
    - Prefix apostrophe if starts with = + - @
    - Cap length to max_len
    """
    if value is None:
        s = ""
    else:
        s = str(value)

    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    cleaned_chars = []
    for ch in s:
        code = ord(ch)
        if code < 32 and code != 32:
            continue
        cleaned_chars.append(ch)
    s = "".join(cleaned_chars)

    if len(s) > max_len:
        s = s[:max_len]

    if s[:1] in {"=", "+", "-", "@"}:
        s = "'" + s
    return s


def freeze_panes(ws: Worksheet, cell: str):
    """
    Freeze panes at a cell (e.g. A2 freezes header row).
    """
    ws.freeze_panes = cell


def autosize_columns_basic(ws: Worksheet, start_col: int, end_col: int, max_width: int = 45):
    """
    Basic autosize by scanning string lengths.
    """
    for c in range(start_col, end_col + 1):
        col_letter = get_column_letter(c)
        best = 10
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            s = str(cell.value)
            best = max(best, min(max_width, len(s) + 2))
        ws.column_dimensions[col_letter].width = best
