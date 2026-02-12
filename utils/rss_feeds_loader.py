from __future__ import annotations

from typing import List, Dict, Any
from openpyxl.workbook.workbook import Workbook


def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"true", "1", "yes", "y", "on"}



def load_rss_feeds_from_wb(wb: Workbook, sheet_name: str = "RSS_Feeds") -> List[Dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, 5)]
    idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    required = ["Source", "URL", "Category", "Active"]
    if not all(r in idx for r in required):
        raise ValueError(f"{sheet_name} missing headers. Need {required}. Found {headers}")

    feeds: List[Dict[str, str]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue

        active = _to_bool(row[idx["Active"]])
        if not active:
            continue

        source = str(row[idx["Source"]] or "").strip()
        url = str(row[idx["URL"]] or "").strip()
        category = str(row[idx["Category"]] or "").strip()

        if not url:
            continue

        feeds.append({"source": source, "url": url, "category": category})

    return feeds
