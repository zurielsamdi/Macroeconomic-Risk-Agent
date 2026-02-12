from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Import your phase functions (we will create these in steps 3–4)
from phase2_curves_ca import run_phase2_curves
from phase4_fx_g10 import run_phase4_fx
from phase5_calendar import run_phase5_calendar
from phase6_headlines import run_phase6_headlines
from phase7_memory_delta import run_phase7_memory_delta
from phase8_predictive import run_phase8_predictive
from phase9_dv01 import run_phase9_dv01
from phase10_fx_mm import run_phase10_fx_mm
from phase11_constraints import run_phase11_constraints
from phase12_commentary import run_phase12_commentary




TEMPLATE_PATH = Path("templates/Morning_Brief_Template.xlsx")
REPORTS_DIR = Path("reports")
LATEST_PATH = REPORTS_DIR / "Morning_Brief_latest.xlsx"


def open_base_workbook():
    """
    If Morning_Brief_latest.xlsx exists, load it (keeps previous phase outputs).
    Otherwise load the template.
    """
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    if LATEST_PATH.exists():
        print(f"Loading existing report: {LATEST_PATH.resolve()}")
        return load_workbook(LATEST_PATH)

    print(f"Loading template: {TEMPLATE_PATH.resolve()}")
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Missing template: {TEMPLATE_PATH.resolve()}")
    return load_workbook(TEMPLATE_PATH)


def main():
    wb = open_base_workbook()

    now = datetime.now()
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")
    file_stamp = now.strftime("%Y%m%d_%H%M%S")

    # Phase 2: Rates / Curves
    rates_end_row = run_phase2_curves(
    wb,
    now_str=now_str,
    file_stamp=file_stamp
)

    # Phase 4: FX
    fx_end_row = run_phase4_fx(wb,now_str=now_str,rates_end_row=rates_end_row)
    calendar_start_row = fx_end_row + 3  # spacing
    calendar_end_row = run_phase5_calendar(wb, start_row=calendar_start_row)
    headlines_start_row = calendar_end_row + 3  # spacing
    headlines_end_row = run_phase6_headlines(wb, now_str=now_str, start_row=headlines_start_row)

    delta_start_row = headlines_end_row + 3
    run_phase7_memory_delta(wb, start_row=delta_start_row)

    # Phase 8-12: predictive, risk engines, constraints, commentary
    try:
        predictive = run_phase8_predictive(wb, now_str=now_str)
    except Exception as e:
        print(f"Phase 8 failed, using conservative fallback: {e}")
        predictive = {
            "rates": {"US": {"rates_vol_risk_label": "medium"}, "CA": {"rates_vol_risk_label": "medium"}},
            "fx": {},
            "events": {"overall_event_severity_label": "medium", "overall_event_severity_score": 3},
            "max_fx_vol_label": "medium",
        }
    try:
        dv01 = run_phase9_dv01(wb, predictive=predictive)
    except Exception as e:
        print(f"Phase 9 failed, using conservative fallback: {e}")
        dv01 = {"net_dv01": 0.0, "front_end_concentration": 0.0, "stress_bp": 2, "stress_$": 0.0}
    try:
        mm = run_phase10_fx_mm(wb, predictive=predictive, now_str=now_str)
    except Exception as e:
        print(f"Phase 10 failed, using conservative fallback: {e}")
        mm = {"inventory_end": 0, "hedge_actions": 0, "avg_spread": 0.00015, "liquidity_risk_label": "medium"}
    try:
        constraints = run_phase11_constraints(predictive=predictive, dv01=dv01, mm=mm)
    except Exception as e:
        print(f"Phase 11 failed, using conservative fallback: {e}")
        constraints = {
            "directional_bias_allowed": False,
            "risk_budget": "low",
            "max_trades": 2,
            "allowed_instruments": [],
            "max_incremental_dv01": 500,
        }
    try:
        run_phase12_commentary(wb, constraints=constraints, now_str=now_str)
    except Exception as e:
        print(f"Phase 12 failed, continuing with workbook save: {e}")

    wb.save(LATEST_PATH)
    print(f"Saved combined report: {LATEST_PATH.resolve()}")

if __name__ == "__main__":
    main()
