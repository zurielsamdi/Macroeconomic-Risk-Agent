from pathlib import Path
from datetime import datetime
import json
from openpyxl import load_workbook

print("RUNNING PHASE0_TEST.PY")

TEMPLATE_PATH = Path("templates/Morning_Brief_Template.xlsx")
OUTPUT_PATH = Path("outputs/Morning_Brief_test.xlsx")
STATE_PATH = Path("state/last_run.json")

def main():
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Missing template: {TEMPLATE_PATH.resolve()}")

    wb = load_workbook(TEMPLATE_PATH)

    sheet_name = "Desk_Commentary"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Template missing sheet '{sheet_name}'. Found: {wb.sheetnames}")

    ws = wb[sheet_name]

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["B1"] = now
    ws["A3"] = "PHASE 0 TEST: Python wrote this successfully."
    ws["A4"] = "If you can see this in Excel, your foundation is working."

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb.save(OUTPUT_PATH)
    STATE_PATH.write_text(json.dumps({"last_run": now}, indent=2), encoding="utf-8")

    print(f"✅ Wrote output file: {OUTPUT_PATH.resolve()}")
    print(f"✅ Wrote state file: {STATE_PATH.resolve()}")

if __name__ == "__main__":
    main()
